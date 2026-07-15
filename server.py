"""
MCQ Hiring-Assessment Localization Tool
English -> Hindi / Marathi / Gujarati, preserving validity as a test item.

Three-layer pipeline (see pipeline.py for full rationale):
    Layer 1: sarvam-translate       -> natural language translation
    Layer 2: glossary + placeholders -> technical term transliteration
    Layer 3: Qwen2.5-7B-Instruct    -> fix MCQ sentence mood + option form

Two models are loaded simultaneously. On an 8GB GPU this requires both to
be 4-bit quantized and is genuinely tight — see README "Memory budget".

Run:
    python server.py
Then open http://localhost:5000
"""

import os
import json
import threading
from flask import Flask, request, jsonify, send_from_directory
from openpyxl import load_workbook, Workbook

from pipeline import (
    get_row_term_candidates,
    build_layer1_request,
    restore_row_terms,
    build_layer3_messages,
)

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────
TRANSLATE_MODEL_PATH = os.environ.get("TRANSLATE_MODEL", r"C:\Users\ASUS\sarvam-translate")
STRUCTURE_MODEL_PATH = os.environ.get("STRUCTURE_MODEL", "Qwen/Qwen2.5-7B-Instruct")
QUANT_MODE = os.environ.get("MODEL_QUANT", "4bit")  # 4bit | 8bit | cpu
MAX_NEW_TOKENS = int(os.environ.get("MAX_NEW_TOKENS", "512"))
PORT = int(os.environ.get("PORT", "5000"))
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "./uploads")
EXPORT_DIR = os.environ.get("EXPORT_DIR", "./exports")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EXPORT_DIR, exist_ok=True)

app = Flask(__name__, static_folder="public")

# ─────────────────────────────────────────────────────────────────────────────
# Model state — TWO models loaded side by side
# ─────────────────────────────────────────────────────────────────────────────
_translate_model = None
_translate_tokenizer = None
_structure_model = None
_structure_tokenizer = None

_load_lock = threading.Lock()
_load_status = "not_started"  # not_started -> loading -> ready -> error
_load_error = None
_device_info = ""


def load_models():
    global _translate_model, _translate_tokenizer, _structure_model, _structure_tokenizer
    global _load_status, _load_error, _device_info

    with _load_lock:
        if _load_status in ("loading", "ready"):
            return
        _load_status = "loading"
        try:
            import torch
            from transformers import AutoModelForCausalLM, AutoTokenizer, BitsAndBytesConfig

            cuda_available = torch.cuda.is_available()
            print(f"[pipeline] CUDA available: {cuda_available}")
            if cuda_available:
                gpu_name = torch.cuda.get_device_name(0)
                vram_gb = torch.cuda.get_device_properties(0).total_memory / 1e9
                _device_info = f"{gpu_name} ({vram_gb:.1f} GB VRAM)"
                print(f"[pipeline] GPU: {_device_info}")
                if vram_gb < 10:
                    print(
                        f"[pipeline] NOTE: {vram_gb:.1f} GB VRAM detected. Running BOTH "
                        "models (sarvam-translate + Qwen2.5-7B) simultaneously in 4-bit "
                        "is tight on an 8GB card. If you hit CUDA OOM, set "
                        "STRUCTURE_MODEL_OFFLOAD=true to load Layer 3's model only "
                        "on-demand and release Layer 1's model from VRAM first — see "
                        "README 'Memory budget' section."
                    )

            target_device = {"": 0} if cuda_available else "cpu"

            def make_bnb_kwargs():
                if QUANT_MODE == "4bit":
                    cfg = BitsAndBytesConfig(
                        load_in_4bit=True,
                        bnb_4bit_quant_type="nf4",
                        bnb_4bit_compute_dtype=torch.float16,
                        bnb_4bit_use_double_quant=True,
                    )
                    return dict(quantization_config=cfg, device_map=target_device)
                elif QUANT_MODE == "8bit":
                    cfg = BitsAndBytesConfig(load_in_8bit=True)
                    return dict(quantization_config=cfg, device_map=target_device)
                elif QUANT_MODE == "cpu":
                    return dict(torch_dtype=torch.float32, device_map="cpu")
                else:
                    return dict(torch_dtype=torch.float16, device_map=target_device)

            print(f"[pipeline] Loading Layer 1 translation model: {TRANSLATE_MODEL_PATH}")
            _translate_tokenizer = AutoTokenizer.from_pretrained(TRANSLATE_MODEL_PATH)
            _translate_model = AutoModelForCausalLM.from_pretrained(
                TRANSLATE_MODEL_PATH, **make_bnb_kwargs()
            )
            print(f"[pipeline] Layer 1 model device: {next(_translate_model.parameters()).device}")

            print(f"[pipeline] Loading Layer 3 structure model: {STRUCTURE_MODEL_PATH}")
            _structure_tokenizer = AutoTokenizer.from_pretrained(STRUCTURE_MODEL_PATH)
            _structure_model = AutoModelForCausalLM.from_pretrained(
                STRUCTURE_MODEL_PATH, **make_bnb_kwargs()
            )
            print(f"[pipeline] Layer 3 model device: {next(_structure_model.parameters()).device}")

            _load_status = "ready"
            print("[pipeline] Both models ready.")
        except Exception as e:
            _load_error = str(e)
            _load_status = "error"
            print(f"[pipeline] FAILED to load models: {e}")


def ensure_loading_started():
    if _load_status == "not_started":
        threading.Thread(target=load_models, daemon=True).start()


ensure_loading_started()


# ─────────────────────────────────────────────────────────────────────────────
# Generation helpers (non-streaming — these run inside a batch pipeline,
# not a chat UI, so we wait for the full result rather than token-stream it)
# ─────────────────────────────────────────────────────────────────────────────
def run_translate(system_prompt: str, user_text: str) -> str:
    import torch
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_text},
    ]
    chat_text = _translate_tokenizer.apply_chat_template(
        messages, tokenize=False, add_generation_prompt=True
    )
    inputs = _translate_tokenizer(chat_text, return_tensors="pt").to(_translate_model.device)
    with torch.no_grad():
        output_ids = _translate_model.generate(
            **inputs,
            max_new_tokens=MAX_NEW_TOKENS,
            do_sample=True,
            temperature=0.01,
            pad_token_id=_translate_tokenizer.eos_token_id,
        )
    new_tokens = output_ids[0][inputs["input_ids"].shape[1]:]
    return _translate_tokenizer.decode(new_tokens, skip_special_tokens=True).strip()


def run_structure_fix(system_prompt: str, user_text: str) -> str:
    import torch
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_text},
    ]
    chat_text = _structure_tokenizer.apply_chat_template(
        messages, tokenize=False, add_generation_prompt=True
    )
    inputs = _structure_tokenizer(chat_text, return_tensors="pt").to(_structure_model.device)
    with torch.no_grad():
        output_ids = _structure_model.generate(
            **inputs,
            max_new_tokens=MAX_NEW_TOKENS,
            do_sample=True,
            temperature=0.2,
            pad_token_id=_structure_tokenizer.eos_token_id,
        )
    new_tokens = output_ids[0][inputs["input_ids"].shape[1]:]
    return _structure_tokenizer.decode(new_tokens, skip_special_tokens=True).strip()


TRANSLATE_LANG_MAP = {"Hindi": "Hindi", "Marathi": "Marathi", "Gujarati": "Gujarati"}


def translate_segment(text: str, language: str) -> str:
    if not text:
        return text
    system = f"Translate the text below to {TRANSLATE_LANG_MAP.get(language, 'Hindi')}."
    return run_translate(system, text)


# ─────────────────────────────────────────────────────────────────────────────
# Routes — Status
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory("public", "index.html")


@app.route("/<path:path>")
def static_files(path):
    return send_from_directory("public", path)


@app.route("/api/status")
def status():
    return jsonify({
        "status": _load_status,
        "translate_model": TRANSLATE_MODEL_PATH,
        "structure_model": STRUCTURE_MODEL_PATH,
        "quant": QUANT_MODE,
        "device": _device_info,
        "error": _load_error,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Routes — File ingestion
# ─────────────────────────────────────────────────────────────────────────────
EXPECTED_COLUMNS = [
    "Sr.No", "Question", "Choice A", "Choice B", "Choice C", "Choice D", "Choice E",
    "Answer", "Type", "Tags", "Row Number", "Allotted Time", "Section", "Hint",
    "User Comment Media", "User Comment Text", "Survey Question(Yes/No)",
    "Data Source (Users/Department/Designation)",
]


@app.route("/api/upload", methods=["POST"])
def upload():

    print("=" * 60)
    print("UPLOAD API CALLED")
    print(request.files)
    print("=" * 60)
    
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    save_path = os.path.join(UPLOAD_DIR, f.filename)
    f.save(save_path)

    wb = load_workbook(save_path, read_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {h: i for i, h in enumerate(header_row) if h}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        def col(name):
            idx = header_map.get(name)
            return r[idx] if idx is not None and idx < len(r) else None

        question = str(col("Question") or "")

        options = [
            str(col("Choice A") or ""),
            str(col("Choice B") or ""),
            str(col("Choice C") or ""),
            str(col("Choice D") or "")
        ]

        candidates = get_row_term_candidates(question, options)
        rows.append({
            "sr_no": col("Sr.No"),
            "question": question,
            "options": options,
            "answer": col("Answer"),
            "type": col("Type"),
            "tags": col("Tags"),
            "section": col("Section"),
            "hint": col("Hint"),
            "candidate_terms": candidates,
        })

    return jsonify({"filename": f.filename, "row_count": len(rows), "rows": rows})


# ─────────────────────────────────────────────────────────────────────────────
# Routes — Pipeline execution (per-row, so the UI can show progress)
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/api/process_row", methods=["POST"])
def process_row():
    """
    Runs all 3 layers for a single row and returns every intermediate
    stage so the review UI can show the full pipeline, not just the
    final result.
    """
    if _load_status != "ready":
        ensure_loading_started()
        return jsonify({"error": f"Models not ready (status: {_load_status}). {_load_error or ''}"}), 503

    data = request.get_json(force=True)
    question = data.get("question", "")
    options = data.get("options", [])
    language = data.get("language", "Hindi")
    confirmed_terms = data.get("confirmed_terms", [])

    # --- Layer 2a: protect confirmed technical terms with placeholders ---
    protected_q, protected_opts, term_map = build_layer1_request(question, options, confirmed_terms)

    # --- Layer 1: translate (placeholders pass through untouched) ---
    translated_q = translate_segment(protected_q, language)
    translated_opts = [translate_segment(o, language) if o else o for o in protected_opts]

    # --- Layer 2b: restore terms as transliterations ---
    restored_q, restored_opts = restore_row_terms(translated_q, translated_opts, term_map, language)

    # --- Layer 3: fix MCQ sentence mood + option form ---
    l3_system, l3_user = build_layer3_messages(language, restored_q, restored_opts)
    raw_l3 = run_structure_fix(l3_system, l3_user)

    final_question, final_options = restored_q, restored_opts
    l3_parse_error = None
    try:
        clean = raw_l3.replace("```json", "").replace("```", "").strip()
        start, end = clean.find("{"), clean.rfind("}")
        parsed = json.loads(clean[start:end + 1])
        final_question = parsed.get("question", restored_q)
        final_options = parsed.get("options", restored_opts)
    except Exception as e:
        l3_parse_error = f"Layer 3 output wasn't valid JSON, falling back to Layer 1+2 result: {e}"

    return jsonify({
        "layer1_2_question": restored_q,
        "layer1_2_options": restored_opts,
        "layer3_raw": raw_l3,
        "final_question": final_question,
        "final_options": final_options,
        "term_map": term_map,
        "l3_parse_error": l3_parse_error,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Routes — Export approved rows to .xlsx matching original schema
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/api/export", methods=["POST"])
def export():
    data = request.get_json(force=True)
    rows = data.get("rows", [])
    language = data.get("language", "Hindi")
    filename = data.get("filename", f"localized_{language.lower()}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Localized {language}"
    ws.append(EXPECTED_COLUMNS)

    for r in rows:
        opts = r.get("final_options", []) + [None] * 5
        ws.append([
            r.get("sr_no"),
            r.get("final_question"),
            opts[0], opts[1], opts[2], opts[3], opts[4],
            r.get("answer"),
            r.get("type"),
            r.get("tags"),
            None, None,
            r.get("section"),
            r.get("hint"),
            None, None, None, None,
        ])

    out_path = os.path.join(EXPORT_DIR, filename)
    wb.save(out_path)
    return jsonify({"path": out_path, "filename": filename})


@app.route("/api/download/<filename>")
def download(filename):
    return send_from_directory(EXPORT_DIR, filename, as_attachment=True)


if __name__ == "__main__":
    print(f"\n{'='*70}")
    print(f"  MCQ Hiring-Assessment Localizer")
    print(f"  Layer 1 (translate): {TRANSLATE_MODEL_PATH}")
    print(f"  Layer 3 (structure): {STRUCTURE_MODEL_PATH}")
    print(f"  Quantization: {QUANT_MODE}")
    print(f"  Open: http://localhost:{PORT}")
    print(f"{'='*70}\n")
    app.run(host="0.0.0.0", port=PORT, threaded=True)
